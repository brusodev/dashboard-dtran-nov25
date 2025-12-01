import openpyxl
from datetime import datetime
import json

print("=" * 100)
print("ANÁLISE DETALHADA DAS PLANILHAS DTRAN-COGESPA")
print("=" * 100)

# ========== PLANILHA 1: FROTA TERCEIRIZADA ==========
print("\n" + "=" * 100)
print("📋 PLANILHA 1: CONTROLE DE TRÁFEGO - FROTA TERCEIRIZADA")
print("=" * 100)

try:
    wb1 = openpyxl.load_workbook('controle de trafego(frota terceirizada) planilha 1.xlsx')
    print(f"\n📑 Abas disponíveis: {wb1.sheetnames}")
    
    for sheet_name in wb1.sheetnames:
        sheet = wb1[sheet_name]
        print(f"\n--- Aba: {sheet_name} ---")
        print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        
        # Mostrar primeiras 15 linhas
        print("\nDados:")
        for row in range(1, min(sheet.max_row + 1, 16)):
            valores = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row, col).value
                valores.append(str(val) if val is not None else "")
            print(f"  Linha {row}: {' | '.join(valores)}")
    
    wb1.close()
except Exception as e:
    print(f"❌ Erro: {e}")

# ========== PLANILHA 2: FROTA INTERNA ==========
print("\n" + "=" * 100)
print("📋 PLANILHA 2: CONTROLE DE TRÁFEGO - FROTA INTERNA")
print("=" * 100)

try:
    wb2 = openpyxl.load_workbook('controle de trafego (frota interna) planilha 2.xlsx')
    print(f"\n📑 Abas disponíveis: {wb2.sheetnames}")
    
    for sheet_name in wb2.sheetnames:
        sheet = wb2[sheet_name]
        print(f"\n--- Aba: {sheet_name} ---")
        print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        
        # Mostrar primeiras 20 linhas
        print("\nDados:")
        for row in range(1, min(sheet.max_row + 1, 21)):
            valores = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row, col).value
                valores.append(str(val) if val is not None else "")
            print(f"  Linha {row}: {' | '.join(valores)}")
    
    wb2.close()
except Exception as e:
    print(f"❌ Erro: {e}")

# ========== PLANILHA 3: MULTAS ==========
print("\n" + "=" * 100)
print("📋 PLANILHA 3: CONTROLE DE MULTAS")
print("=" * 100)

try:
    wb3 = openpyxl.load_workbook('controle de multas planilha 3.xlsx')
    print(f"\n📑 Abas disponíveis: {wb3.sheetnames}")
    
    for sheet_name in wb3.sheetnames:
        sheet = wb3[sheet_name]
        print(f"\n--- Aba: {sheet_name} ---")
        print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        
        # Cabeçalhos
        print("\nCabeçalhos:")
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(1, col).value
            headers.append(header)
            print(f"  Col {col}: {header}")
        
        # Mostrar primeiras 10 linhas de dados
        print("\nPrimeiras 10 linhas de dados:")
        for row in range(2, min(sheet.max_row + 1, 12)):
            valores = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row, col).value
                valores.append(str(val) if val is not None else "")
            print(f"  Linha {row}: {' | '.join(valores)}")
        
        # Análise de diretorias
        print("\n📊 Análise por Diretoria:")
        diretorias = {}
        for row in range(2, sheet.max_row + 1):
            diretoria = sheet.cell(row, 1).value
            if diretoria:
                diretorias[diretoria] = diretorias.get(diretoria, 0) + 1
        
        for dir, count in sorted(diretorias.items(), key=lambda x: x[1], reverse=True):
            print(f"  {dir}: {count} multas")
    
    wb3.close()
except Exception as e:
    print(f"❌ Erro: {e}")

# ========== PLANILHA 4: VEÍCULOS LICENCIADOS ==========
print("\n" + "=" * 100)
print("📋 PLANILHA 4: CONTROLE DE VEÍCULOS LICENCIADOS")
print("=" * 100)

try:
    wb4 = openpyxl.load_workbook('controle de veiculos licenciados planilha 4.xlsx')
    print(f"\n📑 Abas disponíveis: {wb4.sheetnames}")
    
    for sheet_name in wb4.sheetnames:
        sheet = wb4[sheet_name]
        print(f"\n--- Aba: {sheet_name} ---")
        print(f"Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        
        # Mostrar todos os dados
        print("\nTodos os dados:")
        for row in range(1, sheet.max_row + 1):
            valores = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row, col).value
                valores.append(str(val) if val is not None else "")
            print(f"  Linha {row}: {' | '.join(valores)}")
    
    wb4.close()
except Exception as e:
    print(f"❌ Erro: {e}")

print("\n" + "=" * 100)
print("FIM DA ANÁLISE DETALHADA")
print("=" * 100)
